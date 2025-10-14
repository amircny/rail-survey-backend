# backend/main.py
from datetime import datetime
from typing import Any, Dict, List, Optional

from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware

from io import BytesIO
from fastapi.responses import StreamingResponse
import openpyxl
import json
from collections import Counter


from sqlalchemy import (
    Column, Integer, String, JSON, ForeignKey, create_engine, select, delete
)
from sqlalchemy.orm import declarative_base, sessionmaker, Session

# ---------- DB setup ----------
DB_URL = "sqlite:///data.db"
engine = create_engine(DB_URL, echo=False, future=True)
SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
Base = declarative_base()

# ---------- Tables ----------
class Response(Base):
    __tablename__ = "responses"
    id = Column(Integer, primary_key=True, index=True)
    ts = Column(String, nullable=False)
    payload = Column(JSON, nullable=False)

class Question(Base):
    __tablename__ = "questions"
    id = Column(Integer, primary_key=True, index=True)
    text = Column(String, nullable=False)
    qtype = Column(String, nullable=False, default="single")   # single | multi | text
    qorder = Column(Integer, nullable=False, default=0)

class Option(Base):
    __tablename__ = "options"
    id = Column(Integer, primary_key=True, index=True)
    question_id = Column(Integer, ForeignKey("questions.id"), nullable=False)
    code = Column(String, nullable=False)
    label = Column(String, nullable=False)
    oorder = Column(Integer, nullable=False, default=0)

Base.metadata.create_all(engine)

# ---------- App ----------
app = FastAPI(title="Rail Survey API")
app.add_middleware(
    CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"]
)

def db() -> Session:
    s = SessionLocal()
    try:
        yield s
    finally:
        s.close()

# ---------- Health ----------
@app.get("/health")
def health(): return {"ok": True}

# ---------- Responses ----------
@app.post("/submit")
def submit(payload: Dict[str, Any]):
    if not isinstance(payload, dict):
        raise HTTPException(400, "payload must be a JSON object")
    with SessionLocal() as s:
        r = Response(ts=datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"), payload=payload)
        s.add(r); s.commit()
    return {"ok": True}

@app.get("/responses")
def responses() -> List[Dict[str, Any]]:
    with SessionLocal() as s:
        rows = s.execute(select(Response).order_by(Response.id)).scalars().all()
        return [{"id":r.id, "ts":r.ts, "payload":r.payload} for r in rows]

# ---------- Questions CRUD ----------
# ساخت سؤال
@app.post("/question")
def create_question(q: Dict[str, Any]):
    """
    payload نمونه:
    {
      "text": "Which station are you at today?",
      "qtype": "single",
      "qorder": 0,
      "options": [
        {"code":"do","label":"Dortmund Hbf"},
        {"code":"es","label":"Essen Hbf"}
      ]
    }
    """
    text = q.get("text"); qtype = q.get("qtype","single"); qorder = q.get("qorder",0)
    if not text: raise HTTPException(400,"text is required")

    with SessionLocal() as s:
        qrow = Question(text=text, qtype=qtype, qorder=qorder)
        s.add(qrow); s.commit(); s.refresh(qrow)

        opts = q.get("options") or []
        for i, o in enumerate(opts):
            s.add(Option(question_id=qrow.id,
                         code=o.get("code", f"opt{i}"),
                         label=o.get("label",""),
                         oorder=i))
        s.commit()

        return {"id": qrow.id}

# گرفتن همه سؤال‌ها (برای فرانت)
@app.get("/questions")
def get_questions():
    with SessionLocal() as s:
        qrows = s.execute(select(Question).order_by(Question.qorder, Question.id)).scalars().all()
        out = []
        for q in qrows:
            opts = s.execute(
                select(Option).where(Option.question_id==q.id).order_by(Option.oorder, Option.id)
            ).scalars().all()
            out.append({
                "id": q.id, "text": q.text, "type": q.qtype, "order": q.qorder,
                "options": [{"id":o.id,"code":o.code,"label":o.label,"order":o.oorder} for o in opts]
            })
        return {"questions": out}

# ویرایش سؤال (آپدیت کامل + جایگزینی گزینه‌ها)
@app.put("/question/{qid}")
def update_question(qid: int, q: Dict[str, Any]):
    with SessionLocal() as s:
        row = s.get(Question, qid)
        if not row: raise HTTPException(404, "Question not found")

        row.text = q.get("text", row.text)
        row.qtype = q.get("qtype", row.qtype)
        row.qorder = q.get("qorder", row.qorder)
        # حذف گزینه‌های قبلی و ساخت جدید
        s.execute(delete(Option).where(Option.question_id==qid))
        opts = q.get("options") or []
        for i,o in enumerate(opts):
            s.add(Option(question_id=qid, code=o.get("code",f"opt{i}"), label=o.get("label",""), oorder=i))
        s.add(row); s.commit()
        return {"ok": True}

# حذف سؤال
@app.delete("/question/{qid}")
def delete_question(qid: int):
    with SessionLocal() as s:
        s.execute(delete(Option).where(Option.question_id==qid))
        row = s.get(Question, qid)
        if row: s.delete(row)
        s.commit()
    return {"ok": True}
# ---------- Excel Export ----------
from io import BytesIO
from fastapi.responses import StreamingResponse
import openpyxl, json
from collections import Counter
from sqlalchemy import select

@app.get("/export.xlsx")
def export_excel():
    # 1) داده‌ها از DB
    with SessionLocal() as s:
        qrows = s.execute(select(Question).order_by(Question.qorder, Question.id)).scalars().all()
        orows = s.execute(select(Option).order_by(Option.oorder, Option.id)).scalars().all()
        rrows = s.execute(select(Response).order_by(Response.id)).scalars().all()

    # 2) map گزینه‌ها: {qid: {code: label}}
    optmap = {}
    for o in orows:
        optmap.setdefault(o.question_id, {})[o.code] = o.label

    # 3) ساخت Workbook
    wb = openpyxl.Workbook()

    # Sheet خام
    ws_raw = wb.active
    ws_raw.title = "raw_responses"
    ws_raw.append(["id", "ts", "payload"])
    for r in rrows:
        ws_raw.append([r.id, r.ts, json.dumps(r.payload, ensure_ascii=False)])

    # Sheet خلاصه کل
    ws_sum = wb.create_sheet("summary_counts")
    ws_sum.append(["Question", "Option", "Count"])

    # شیت برای هر سؤال
    for q in qrows:
        qid, qtext, qtype = q.id, q.text, q.qtype

        if qtype in ("single", "multi"):
            counts = Counter()
            for r in rrows:
                ans = (r.payload or {}).get("answers", {})
                v = ans.get(str(qid)) if isinstance(ans, dict) else None
                if v is None:
                    continue
                if qtype == "multi" and isinstance(v, list):
                    for code in v:
                        counts[code] += 1
                else:
                    counts[v] += 1

            ws = wb.create_sheet(f"Q{qid}_counts")
            ws.append(["Question", "Option", "Count"])
            for code, n in counts.items():
                label = optmap.get(qid, {}).get(code, str(code))
                ws.append([qtext, label, n])
                ws_sum.append([qtext, label, n])

        else:  # text
            ws = wb.create_sheet(f"Q{qid}_texts")
            ws.append(["response_id", "ts", "Answer"])
            for r in rrows:
                ans = (r.payload or {}).get("answers", {})
                v = ans.get(str(qid)) if isinstance(ans, dict) else None
                if v is not None:
                    ws.append([r.id, r.ts, str(v)])

    # 4) ارسال فایل به‌صورت استریم
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    headers = {"Content-Disposition": 'attachment; filename="survey_export.xlsx"'}
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
# ---------- Excel Export (FLAT) ----------
from io import BytesIO
from fastapi.responses import StreamingResponse
import openpyxl, json
from sqlalchemy import select

@app.get("/export_flat.xlsx")
def export_excel_flat():
    # 1) load data
    with SessionLocal() as s:
        qrows = s.execute(select(Question).order_by(Question.qorder, Question.id)).scalars().all()
        orows = s.execute(select(Option).order_by(Option.oorder, Option.id)).scalars().all()
        rrows = s.execute(select(Response).order_by(Response.id)).scalars().all()

    # map: question id -> {code: label}
    optmap = {}
    for o in orows:
        optmap.setdefault(o.question_id, {})[o.code] = o.label

    # 2) prepare header: ts + each question text (in order)
    header = ["ts"] + [q.text for q in qrows]

    # 3) build rows: each response -> one row of human-readable labels
    rows = []
    for r in rrows:
        ans = (r.payload or {}).get("answers", {})
        row = [r.ts]
        for q in qrows:
            qid, qtype = q.id, q.qtype
            v = ans.get(str(qid))
            if v is None:
                row.append("")  # no answer
            elif qtype == "text":
                row.append(str(v))
            elif qtype == "multi":
            # multi choice -> just save English codes
              row.append(", ".join(v or []))

            else:  # single choice -> save English code
             row.append(v)
        rows.append(row)

    # 4) write to Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "flat"
    ws.append(header)
    for row in rows:
        ws.append(row)

    # autosize columns (basic)
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(60, max(10, max_len + 2))

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="survey_export_flat.xlsx"'},
    )

